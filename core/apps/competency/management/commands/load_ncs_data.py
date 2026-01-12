import os
from argparse import ArgumentParser
from collections.abc import Iterable
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils.translation import gettext as _

from apps.competency.models import Classification, Factor, Skill


class Command(BaseCommand):
    help = _("Import classifications, competency skills, and skill factors from Excel file")

    def add_arguments(self, parser: ArgumentParser):
        parser.add_argument(
            "--file",
            type=str,
            default="apps/competency/fixtures/ncs-241204.xlsx",
            help=_("Path to the competency data excel file"),
        )

    @transaction.atomic
    def handle(self, *args: Iterable[object], **options: str):
        file_path = options["file"]
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(_("File not found: %s") % file_path))
            return

        self.stdout.write(self.style.SUCCESS(_("Processing file: %s") % file_path))
        self._import_data(file_path)
        self.stdout.write(self.style.SUCCESS(_("Import completed successfully")))

    def _import_data(self, file_path: str):
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        version = Path(file_path).stem

        self._import_classifications(workbook, version)
        self._import_competency_skills_and_factors(workbook, version)

    def _import_classifications(self, workbook: openpyxl.Workbook, version: str):
        sheet = workbook.worksheets[0]

        code_name_dict: dict[str, str] = {}
        parent_dict: dict[str, str | None] = {}
        level_dict: dict[str, int] = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            large_code = str(row[0]).strip()
            large_name = str(row[1]).strip()
            medium_code = str(row[2]).strip()
            medium_name = str(row[3]).strip()
            small_code = str(row[4]).strip()
            small_name = str(row[5]).strip()
            detail_code = str(row[6]).strip() if row[6] else None
            detail_name = str(row[7]).strip() if row[7] else None

            large_full_code = large_code
            medium_full_code = f"{large_code}{medium_code}"
            small_full_code = f"{medium_full_code}{small_code}"
            detail_full_code = f"{small_full_code}{detail_code}" if detail_code else None

            code_name_dict[large_full_code] = large_name
            level_dict[large_full_code] = 1

            code_name_dict[medium_full_code] = medium_name
            level_dict[medium_full_code] = 2

            code_name_dict[small_full_code] = small_name
            level_dict[small_full_code] = 3

            if detail_full_code and detail_name:
                code_name_dict[detail_full_code] = detail_name
                level_dict[detail_full_code] = 4

            parent_dict[medium_full_code] = large_full_code
            parent_dict[small_full_code] = medium_full_code
            if detail_full_code:
                parent_dict[detail_full_code] = small_full_code

        existing_classifications = {c.code: c for c in Classification.objects.all()}

        for current_level in [1, 2, 3, 4]:
            for code, name in code_name_dict.items():
                if level_dict.get(code, 0) != current_level:
                    continue

                if code in existing_classifications:
                    if existing_classifications[code].name != name:
                        existing_classifications[code].name = name
                        existing_classifications[code].version = version
                        existing_classifications[code].save()
                    continue

                parent_code = parent_dict.get(code)
                if current_level == 1:
                    new_node = Classification.add_root(code=code, name=name, version=version)
                else:
                    if not parent_code or parent_code not in existing_classifications:
                        continue

                    parent = existing_classifications[parent_code]
                    new_node = parent.add_child(code=code, name=name, version=version)

                existing_classifications[code] = new_node

        Classification.fix_tree()

    def _import_competency_skills_and_factors(self, workbook: openpyxl.Workbook, version: str):
        competency_sheet = workbook.worksheets[0]

        classification_lookup = {c.code: c for c in Classification.objects.all()}
        competency_skills_data: list[Skill] = []
        competency_skill_keys: set[str] = set()

        for row in competency_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[8]:
                continue

            large_code = str(row[0]).strip()
            medium_code = str(row[2]).strip()
            small_code = str(row[4]).strip()
            detail_code = str(row[6]).strip() if row[6] else ""

            if detail_code:
                classification_code = f"{large_code}{medium_code}{small_code}{detail_code}"
            else:
                classification_code = f"{large_code}{medium_code}{small_code}"

            if classification_code not in classification_lookup:
                self.stdout.write(self.style.WARNING(f"Classification not found: {classification_code}"))
                continue

            skill_number = str(row[8]).strip()
            skill_code = str(row[9]).strip()
            skill_name = str(row[10]).strip()
            skill_level = int(str(row[11])) if row[11] else 1

            skill_key = skill_number
            if skill_key in competency_skill_keys:
                continue

            competency_skill_keys.add(skill_key)

            competency_skills_data.append(
                Skill(
                    classification=classification_lookup[classification_code],
                    code=skill_code,
                    name=skill_name,
                    level=skill_level,
                    number=skill_number,
                    version=version,
                )
            )

        self._bulk_update_or_create_competency_skills(competency_skills_data)

        self._import_skill_factors(workbook, version)

    def _bulk_update_or_create_competency_skills(self, skills_data: list[Skill]):
        existing_skills = {skill.number: skill for skill in Skill.objects.all()}

        skills_to_create: list[Skill] = []
        skills_to_update: list[Skill] = []

        for skill_data in skills_data:
            number = skill_data.number

            if number in existing_skills:
                skill = existing_skills[number]
                changed = False

                if skill.code != skill_data.code:
                    skill.code = skill_data.code
                    changed = True
                if skill.name != skill_data.name:
                    skill.name = skill_data.name
                    changed = True
                if skill.level != skill_data.level:
                    skill.level = skill_data.level
                    changed = True
                if skill.classification_id != skill_data.classification_id:
                    skill.classification = skill_data.classification
                    changed = True
                if skill.version != skill_data.version:
                    skill.version = skill_data.version
                    changed = True
                if changed:
                    skills_to_update.append(skill)
            else:
                skills_to_create.append(
                    Skill(
                        classification=skill_data.classification,
                        code=skill_data.code,
                        name=skill_data.name,
                        level=skill_data.level,
                        number=skill_data.number,
                        version=skill_data.version,
                    )
                )

        if skills_to_create:
            Skill.objects.bulk_create(skills_to_create)

        if skills_to_update:
            Skill.objects.bulk_update(skills_to_update, ["code", "name", "level", "classification_id", "version"])

    def _import_skill_factors(self, workbook: openpyxl.Workbook, version: str):
        factor_sheet = workbook.worksheets[1]

        skill_lookup = {skill.number: skill for skill in Skill.objects.all()}
        factors_data: list[Factor] = []
        factor_keys: set[str] = set()

        for row in factor_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[8]:
                continue

            skill_number = str(row[8]).strip()
            if skill_number not in skill_lookup:
                self.stdout.write(self.style.WARNING(f"Skill not found: {skill_number}"))
                continue

            factor_key = str(row[12])
            if factor_key in factor_keys:
                continue

            factor_keys.add(factor_key)

            factors_data.append(
                Factor(
                    skill=skill_lookup[skill_number], code=str(row[11]), name=row[13], number=row[12], version=version
                )
            )

        self._bulk_update_or_create_skill_factors(factors_data)

    def _bulk_update_or_create_skill_factors(self, factors_data: list[Factor]):
        existing_factors = {factor.number: factor for factor in Factor.objects.all()}

        factors_to_create: list[Factor] = []
        factors_to_update: list[Factor] = []

        for factor_data in factors_data:
            number = factor_data.number

            if number in existing_factors:
                factor = existing_factors[number]
                changed = False

                if factor.code != factor_data.code:
                    factor.code = factor_data.code
                    changed = True
                if factor.name != factor_data.name:
                    factor.name = factor_data.name
                    changed = True
                if factor.skill_id != factor_data.skill_id:
                    factor.skill = factor_data.skill
                    changed = True
                if factor.version != factor_data.version:
                    factor.version = factor_data.version
                    changed = True

                if changed:
                    factors_to_update.append(factor)
            else:
                factors_to_create.append(
                    Factor(
                        skill=factor_data.skill,
                        code=factor_data.code,
                        name=factor_data.name,
                        number=factor_data.number,
                        version=factor_data.version,
                    )
                )

        if factors_to_create:
            Factor.objects.bulk_create(factors_to_create)

        if factors_to_update:
            Factor.objects.bulk_update(factors_to_update, ["code", "name", "skill_id", "version"])
